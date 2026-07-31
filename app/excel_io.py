"""
Excel/CSV input reading and Excel output writing.

Output layout: one row per postal code. Row 1: location name, merged across that
location's 6 columns. Row 2: column headers (Temporary ID, Postal code, then per
location: drive_distance_km, transit_distance_km, flyover_distance_km, drive_duration,
transit_duration, flyover_duration). Row 3+: data. Durations are minutes, distances are
km. A blank cell means that value was unavailable (no route found), or -- flyover only --
not applicable (origin and destination share the same nearest airport).
"""

import io
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .pipeline import LOCATION_FIELDS

POSTAL_CODE_COLUMN_ALIASES = {"postalcode", "postal_code", "postal code", "pc", "postal"}
TEMPORARY_ID_COLUMN_ALIASES = {"temporary id", "temporary_id", "temp id", "temp_id", "tempid", "id"}


def _normalize_col(name: str) -> str:
    return str(name).strip().lower().replace("_", " ")


def _find_column(df: pd.DataFrame, aliases: set[str]) -> str | None:
    for c in df.columns:
        if _normalize_col(c) in aliases:
            return c
    return None


def _coerce_id(value) -> int | str | None:
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return text


def _extract_rows(df: pd.DataFrame, source_name: str) -> list[dict]:
    """Returns [{"temporary_id": ..., "postal_code": ...}, ...]. If the input has its own
    Temporary ID column, that value is carried through to the output as-is; otherwise rows
    are numbered sequentially (1, 2, 3, ...) based on their position in the input file."""
    if df.shape[1] == 1:
        pc_column = df.columns[0]
        id_column = None
    else:
        pc_column = _find_column(df, POSTAL_CODE_COLUMN_ALIASES)
        if pc_column is None:
            raise ValueError(
                f"Could not find a postal code column in {source_name}. "
                f"Expected a column named one of {sorted(POSTAL_CODE_COLUMN_ALIASES)}, "
                f"or a single-column file. Found columns: {list(df.columns)}"
            )
        id_column = _find_column(df, TEMPORARY_ID_COLUMN_ALIASES)

    rows = []
    for position, (_, record) in enumerate(df.iterrows(), start=1):
        postal_code = str(record[pc_column]).strip() if pd.notna(record[pc_column]) else ""
        if not postal_code:
            continue
        temporary_id = _coerce_id(record[id_column]) if id_column is not None else None
        rows.append({"temporary_id": temporary_id if temporary_id is not None else position, "postal_code": postal_code})
    return rows


def _read_dataframe(source, suffix: str) -> pd.DataFrame:
    suffix = suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(source, dtype=str)
    if suffix in (".xlsx", ".xls"):
        return pd.read_excel(source, dtype=str)
    raise ValueError(f"Unsupported input file type: {suffix} (expected .csv, .xlsx, or .xls)")


def read_input_rows_from_buffer(buffer, filename: str) -> list[dict]:
    """Reads an uploaded .csv/.xlsx (in-memory) and returns
    [{"temporary_id": ..., "postal_code": ...}, ...], one per input row."""
    df = _read_dataframe(buffer, Path(filename).suffix)
    return _extract_rows(df, filename)


def build_workbook(rows: list[dict]) -> openpyxl.Workbook:
    """rows: [{"temporary_id": int, "postal_code": str, "locations": [...]}, ...]
    (the dicts returned by pipeline.process_postal_code, plus a "temporary_id" key)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    location_names = [loc["name"] for loc in rows[0]["locations"]] if rows else []

    ws.cell(row=2, column=1, value="Temporary ID")
    ws.cell(row=2, column=2, value="Postal code")

    col = 3
    for name in location_names:
        start_col = col
        for field in LOCATION_FIELDS:
            ws.cell(row=2, column=col, value=field)
            col += 1
        ws.cell(row=1, column=start_col, value=name)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col - 1)
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center")

    last_col = col - 1
    for r, row in enumerate(rows, start=3):
        ws.cell(row=r, column=1, value=row.get("temporary_id"))
        ws.cell(row=r, column=2, value=row.get("postal_code"))
        col = 3
        for loc in row["locations"]:
            for field in LOCATION_FIELDS:
                ws.cell(row=r, column=col, value=loc.get(field))
                col += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    for i in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    return wb


def build_workbook_bytes(rows: list[dict]) -> bytes:
    """Workbook as in-memory bytes -- the FastAPI batch endpoint returns this directly in
    the HTTP response, without touching disk."""
    buffer = io.BytesIO()
    build_workbook(rows).save(buffer)
    return buffer.getvalue()
